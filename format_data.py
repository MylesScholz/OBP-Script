"""
Created on May 14th, 2018
Author: Miles McCall
Modified by Myles Scholz, June 2023
Sources:
Description: Parse pollinator data pulled from iNaturalist.org and generate a reformatted version.
"""

# External imports
import os
import sys
import string
import errno
import csv
import openpyxl
from openpyxl import load_workbook
from openpyxl import Workbook

# Local imports
import col_functions


# Functions


def letter_to_index(letter):
    """Converts a column letter, e.g. "A", "B", "AA", "BC" etc. to a zero based
    column index.
    A becomes 0, B becomes 1, Z becomes 25, AA becomes 26 etc.

    Args:
        letter (str): The column index letter.

    Returns:
        The column index as an integer.
    """

    letter = letter.upper()

    result = 0

    for index, char in enumerate(reversed(letter)):
        # Get the ASCII number of the letter and subtract 64 so that A
        # corresponds to 1.

        num = ord(char) - 64

        # Multiply the number with 26 to the power of `index` to get the correct
        # value of the letter based on it's index in the string.

        final_num = (26**index) * num

        result += final_num

    # Subtract 1 from the result to make it zero-based before returning.

    return result - 1


def count_rows(workbook, worksheet):
    wb = load_workbook(str(workbook))

    ws = wb[str(worksheet)]

    row_count = str(ws.max_row)

    print("count rows:", row_count)

    return row_count


def count_cols(workbook, worksheet):
    wb = load_workbook(str(workbook))

    ws = wb[str(worksheet)]

    col_count = str(ws.max_column)

    print("count cols:", col_count)

    return col_count


def parse_cmd_line():
    # Parse the command line arguments:
    # Determine an input path, output path, and input file type

    print("Parsing command line arguments...")

    # Init vars
    i = 0
    input_path = ""
    output_path = ""
    input_file_type = ""

    # Input file is required
    if "--input" not in sys.argv:
        print("ERROR: --input argument must be provided")
        exit(1)

    # Iterate through command line and assign strings to input and output paths
    for arg in sys.argv:
        if arg == "--input":
            if i + 1 >= len(sys.argv):
                print("ERROR: --input argument not set")
                exit(1)

            input_path = sys.argv[i + 1]
        elif arg == "--output":
            if i + 1 >= len(sys.argv):
                print("ERROR: --output argument not set")
                exit(1)

            output_path = sys.argv[i + 1]

        i += 1

    # Input path:
    # data/folder_name/file_name

    # The input file must be kept in data dir
    if input_path.split("/")[0] != "data":
        print("ERROR: --input file must be saved inside the 'data' directory")
        exit(1)

    # Parse input file and input file type
    input_path_split = input_path.split("/")
    input_file = input_path_split[-1]

    input_file_split = input_file.split(".")
    input_file_name = input_file_split[0]
    input_file_type = input_file_split[1]

    # Output path:
    # results/folder_name/file_name

    # If the output file does exist, it must be kept in the results dir
    if output_path != "" and output_path.split("/")[0] != "results":
        print("ERROR: --output file must be saved inside the 'results' directory")
        exit(1)

    # If output was not specified, use the input folder name
    if output_path == "":
        # We will use the split file path components
        output_path = "results/" + input_path.split("/")[1] + "/results.csv"

    # Remove file from output path to get the output directory
    # Append '/' so it is treated as a directory
    output_folder = output_path[: output_path.rindex("/")] + "/"

    # Create output folder
    if not os.path.exists(os.path.dirname(output_folder)):
        try:
            os.makedirs(os.path.dirname(output_folder))
        except OSError as exc:  # Guard against race condition
            if exc.errno != errno.EEXIST:
                raise

    # Create output file
    f = open(output_path, "w")
    f.close()

    return (
        input_path,
        input_file_name,
        input_file_type.lower(),
        output_path,
    )


def read_csv_header(file_string):
    file = open(file_string, "r")  # Open CSV file

    return file.readline()


def read_xlsx_header(wb_name, ws_name):
    wb = load_workbook(wb_name)

    ws = wb[str(ws_name)]

    max_col = list(string.ascii_lowercase)[ws.max_column - 1].upper()

    print(max_col)

    print(ws.max_column)

    selection_str = "A1:" + max_col + "1"

    return ws[selection_str].ecnode("utf8")


def read_csv(file_string):
    print("\tReading from CSV...")

    # Open CSV file
    file = open(file_string, "r")

    # Read the first line into header var
    header_row = file.readline()

    # Iterate through rest of file, saving in array
    file_rows = []

    for line in file:
        file_rows.append(line)

    # Returns the header row and line array
    return header_row, file_rows


def read_xlsx(file_string):
    print("\tReading from Excel spreadsheet...")
    # TODO: Read xlsx header and content, save them to strings, and return them


def read_data(file_string, file_type):
    print("Reading data from input source...")

    # Variables to capture the header row and following data
    header = ""
    file_rows = []

    # Check which file type to read from
    if file_type == "csv":
        header, file_rows = read_csv(file_string)
    elif file_type == "xlsx":
        header, file_rows = read_xlsx(file_string)
    else:
        print("ERROR: Invalid input file type")
        exit(1)

    # Strip header of extra characters and converts individuals chars into words
    header = header.strip().split(",")

    return header, file_rows


def write_list_to_csv(list: list, csv_file: str):
    with open(csv_file, "a", newline="") as file:
        writer = csv.writer(file, delimiter=",", lineterminator="\n")
        writer.writerow(list)


def write_string_to_csv(line: str, csv_file: str):
    with open(csv_file, "a", newline="") as file:
        writer = csv.writer(file, delimiter=",", lineterminator="\n")
        writer.writerow(line.split(","))


def remove_blank_rows(out_file):
    header, rows = read_csv(out_file)

    write_list_to_csv(header, out_file)

    for line in rows:
        if line != "\n":
            write_string_to_csv(line, out_file)


def get_row_value_by_column(header: list, row: list, column_name: str):
    try:
        column_index = header.index(column_name)
        return row[column_index]
    except ValueError:
        # column_name could not be found in header list
        return ""


def print_over_line(string: str):
    print("\033[K" + string + "\r", end="")


def gen_output(
    output_header: list, output_file: str, input_header: list, input_rows: list
):
    """
    Generate data for the output file

    :param output_header: list of column labels (the header) for the output CSV
    :param output_file: output file path
    :param input_header: list of column labels (the header) for the input CSV
    :param input_rows: list of lines (strings) read from the input CSV
    """

    # Create rows of formatted data and append to output csv
    print("Generating output data...")
    # Print header row
    write_list_to_csv(output_header, output_file)

    # Counter for printing out
    row_count = 0

    # Parse input rows
    # csv.reader automatically splits input_row on commas into a list of strings
    for input_row in csv.reader(input_rows, skipinitialspace=True):
        # Init the output row
        output_row = []
        # Observation No.
        # Voucher No.
        output_row.append(" ")
        output_row.append(" ")

        # iNaturalist ID
        iNaturalist_id = get_row_value_by_column(input_header, input_row, "user_id")
        output_row.append(iNaturalist_id)

        # iNaturalist Alias
        iNaturalist_alias = get_row_value_by_column(
            input_header, input_row, "user_login"
        )
        output_row.append(iNaturalist_alias)

        # Collector - First Name
        # Collector - First Initial
        # Collector - Last Name
        user_full_name = get_row_value_by_column(
            input_header, input_row, "user_full_name"
        )
        (
            user_first_name,
            user_first_initial,
            user_last_name,
        ) = col_functions.split_user_name(iNaturalist_alias, user_full_name)

        output_row.append(user_first_name)
        output_row.append(user_first_initial)
        output_row.append(user_last_name)

        # Sample ID
        sample_id = get_row_value_by_column(input_header, input_row, "field:sample id.")
        output_row.append(sample_id)

        # Specimen ID
        specimen_id = get_row_value_by_column(
            input_header, input_row, "field:number of bees collected"
        )
        output_row.append(specimen_id)

        # Collection Day 1
        # Month 1
        # Year 1
        # Time 1
        date1 = get_row_value_by_column(input_header, input_row, "observed_on")
        day1, month1, year1 = col_functions.date_1(date1)
        time1 = col_functions.time_1(input_row[input_header.index("time_observed_at")])
        output_row.append(day1)
        output_row.append(month1)
        output_row.append(year1)
        output_row.append(time1)

        # Collection Day 2
        # Month 2
        # Year 2
        # Time 2
        date2 = get_row_value_by_column(input_header, input_row, "field:trap removed")
        day2, month2, year2, merge2 = col_functions.date_2(date2)
        time2 = col_functions.time_2(date2)
        output_row.append(day2)
        output_row.append(month2)
        output_row.append(year2)
        output_row.append(time2)

        # Country
        country = "USA"
        output_row.append(country)

        # State
        state = get_row_value_by_column(input_header, input_row, "place_state_name")
        if state == "Oregon":
            state = "OR"
        output_row.append(state)

        # County
        county = get_row_value_by_column(input_header, input_row, "place_county_name")
        output_row.append(county)

        # Location
        place_guess = get_row_value_by_column(input_header, input_row, "place_guess")
        location = col_functions.location_guess(place_guess)
        output_row.append(location)

        # Site Description
        site_description = get_row_value_by_column(
            input_header, input_row, "field:collection site description"
        )
        output_row.append(site_description)

        # Abbreviated Location
        abbreviated_location = location
        output_row.append(abbreviated_location)

        # Dec. Lat.
        # Dec. Long.
        lat = get_row_value_by_column(input_header, input_row, "latitude")
        long = get_row_value_by_column(input_header, input_row, "longitude")
        if lat == "" or long == "":
            output_row.append("")
            output_row.append("")
        else:
            lat = col_functions.round_coord(lat)
            long = col_functions.round_coord(long)
            if lat is None or long is None:
                output_row.append("")
                output_row.append("")
            else:
                output_row.append(lat)
                output_row.append(long)

        # Pos Accuracy
        pos_acc = get_row_value_by_column(
            input_header, input_row, "positional_accuracy"
        )
        output_row.append(pos_acc)

        # Elevation
        if lat is None or long is None or lat == "" or long == "":
            output_row.append("")
        else:
            elevation = col_functions.elevation(lat, long)
            output_row.append(elevation)

        # Collection method
        # collection_method = check_for_cols(in_header, in_row, "field:oba collection method")
        # only use first word of collection method
        collection_string = get_row_value_by_column(
            input_header, input_row, "field:oba collection method"
        )
        collection_method = col_functions.collection(collection_string)
        output_row.append(collection_method)

        # Associated plant - family
        # Associated plant - species
        # Associated plant - iNaturalist url
        family = get_row_value_by_column(input_header, input_row, "taxon_family_name")
        species = get_row_value_by_column(input_header, input_row, "scientific_name")
        url = get_row_value_by_column(input_header, input_row, "url")
        output_row.append(family)
        output_row.append(species)
        output_row.append(url)

        # End of appending to output row

        # Append generated row to output file
        # If the row has multiple bees collected, expand by that many
        if specimen_id is not None and specimen_id != "":
            try:
                specimen_id = int(specimen_id)
                if int(specimen_id) >= 1:
                    # print("Multiple bees, printing", specimenid, "times...")
                    for i in range(1, int(specimen_id) + 1):
                        output_row[output_header.index("Specimen ID")] = i
                        write_list_to_csv(output_row, output_file)
                        # print(out_row)
            except ValueError:
                # it was a string, not an int.
                write_list_to_csv(output_row, output_file)
                # print(out_row)

            row_count += 1
            print_over_line("\t" + str(row_count))

    print()


def main():
    # Intro
    print("iNaturalist Pipeline -----------------------")

    # I/O variables
    input_file = ""
    output_file = ""
    input_file_type = ""

    # Parse command line arguments
    (
        input_file,
        input_file_name,
        input_file_type,
        output_file,
    ) = parse_cmd_line()

    # Pipeline Description
    print("\tInput path:\t", input_file)
    print("\tInput name:\t", input_file_name)
    print("\tInput type:\t", input_file_type)
    print("\tOutput path:\t", output_file)
    print()

    # Choose which file reading function to call
    input_header, input_rows = read_data(input_file, input_file_type)

    print()
    # Sort columns before writing output
    output_header = [
        "Observation No.",
        "Voucher No.",
        "iNaturalist ID",
        "iNaturalist Alias",
        "Collector - First Name",
        "Collector - First Name Initial",
        "Collector - Last Name",
        "Sample ID",
        "Specimen ID",
        "Collection Day 1",
        "Month 1",
        "Year 1",
        "Time 1",
        "Collection Day 2",
        "Month 2",
        "Year 2",
        "Time 2",
        "Country",
        "State",
        "County",
        "Location",
        "Collection Site Description",
        "Abbreviated Location",
        "Dec. Lat.",
        "Dec. Long.",
        "Lat/Long Accuracy",
        "Elevation",
        "Collection method",
        "Associated plant - family",
        "Associated plant - species",
        "Associated plant - Inaturalist URL",
    ]

    # Create output data
    gen_output(output_header, output_file, input_header, input_rows)

    print()


if __name__ == "__main__":
    main()
